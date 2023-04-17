
from tkinter import *
from tkinter import ttk
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl,xlrd
from openpyxl import workbook
import pathlib 




root=Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False,False)
root.configure(bg="#326273")


file=pathlib.Path('Backdata.xlsx')
if file.exists():
    pass
else:
    workbook=openpyxl.Workbook()
    sheet=workbook.active
    sheet['A1']="Full Name"
    sheet['B1']="Phone muber"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"

    workbook.save(file)


def Submit():
    name=namevalue.get()
    contact=contactvalue.get()
    age=Agevalue.get()
    gender=gender_combobox.get()
    address=addressEntry.get(1.0,END)

    file=openpyxl.load_workbook('Backdata.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)

    file.save(r'Backdata.xlsx')

    messagebox.showinfo('info','Submitted Successfully !')
    
    namevalue.set('')
    contactvalue.set('')
    Agevalue,set('')
    addressEntry.delete(1.0,END)



def clear():
    namevalue.set('')
    contactvalue.set('')
    Agevalue,set('')
    addressEntry.delete(1.0,END)

#icon
root.iconbitmap(r'C:\Users\hp\Desktop\python project\logo.ico')

#heading
Label(root,text="Please fill out this form:",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)

#label
Label(root,text='Name',font=23,bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text='Contact No.',font=23,bg="#326273",fg="#fff").place(x=50,y=150)
Label(root,text='Age',font=23,bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text='Gender',font=23,bg="#326273",fg="#fff").place(x=300,y=200)
Label(root,text='Address',font=23,bg="#326273",fg="#fff").place(x=50,y=250)


#Entry
namevalue=StringVar()
contactvalue=StringVar()
Agevalue=StringVar()

nameEntry=Entry(root,textvariable=namevalue,width=30, bd=2,font=18)
contactEntry=Entry(root,textvariable=contactvalue,width=30, bd=2,font=18)
ageEntry=Entry(root,textvariable=Agevalue,width=5, bd=2,font=18)


#gender
gender_combobox=Combobox(root,values=['Male','Female'],font='arial 14',state='r',width=10)
gender_combobox.place(x=400,y=200)


#address

addressEntry=Text(root,width=40,height=4,bd=4)



nameEntry.place(x=200,y=100)
contactEntry.place(x=200,y=150)
ageEntry.place(x=200,y=200)
addressEntry.place(x=200,y=250)

Button(root,text="Submit",bg="#326275",fg="white",width=15,height=2,command=Submit).place(x=200,y=350)
Button(root,text="Clear",bg="#326263",fg="white",width=15,height=2,command=clear).place(x=340,y=350)
Button(root,text="Exit",bg="#326200",fg="white",width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)


root.mainloop()
